"""Per-school roster xlsx generator.

For each school we byte-copy the bundled template, then open the copy with
openpyxl and write three cell values per delegate row (School / Committee /
Position). The literal file copy guarantees we start from an identical file;
openpyxl only re-serializes on save and only writes cell values, so merges,
fonts, column widths, and row heights all survive.

openpyxl does NOT round-trip embedded images, so the secretariat's banner
image gets dropped on save. We compensate by bundling the raw image
alongside the template and re-attaching it on the worksheet at the same
anchor (column C, row 0) and dimensions as the original drawing.
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from whsmun.committees import COMMITTEES, ALL_COMMITTEES, WEIGHT
from whsmun.reporting.csv_reader import SchoolRoster

_KIND = {c.canonical: c.kind for c in COMMITTEES}
_FIRST_DATA_ROW = 7
_INVALID_FILENAME_CHARS = '/\\:*?"<>|'
# Banner image position & size mirror the original template's drawing1.xml
# (oneCellAnchor at col=2 row=0, ext cx=2209800 cy=2171700 in EMUs).
_IMAGE_ANCHOR = "C1"
_EMU_PER_PIXEL = 9525
_IMAGE_WIDTH_PX = 2209800 / _EMU_PER_PIXEL
_IMAGE_HEIGHT_PX = 2171700 / _EMU_PER_PIXEL


def write_rosters(
    template_path: Path,
    output_dir: Path,
    rosters: Iterable[SchoolRoster],
    image_path: Path | None = None,
) -> int:
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True)
    count = 0
    for roster in rosters:
        _write_one(template_path, output_dir, roster, image_path)
        count += 1
    return count


def _write_one(
    template_path: Path,
    output_dir: Path,
    roster: SchoolRoster,
    image_path: Path | None,
) -> Path:
    out_path = output_dir / f"{_sanitize(roster.name)} WHSMUN Roster.xlsx"
    shutil.copy2(template_path, out_path)
    wb = load_workbook(out_path)
    ws = wb["ROSTER TEMPLATE"]
    row = _FIRST_DATA_ROW
    for committee in ALL_COMMITTEES:
        count = roster.placements.get(committee, 0)
        if count == 0:
            continue
        for school, committee_name, position in _rows_for(roster, committee, count):
            ws.cell(row=row, column=1, value=school)
            ws.cell(row=row, column=2, value=committee_name)
            if position is not None:
                ws.cell(row=row, column=3, value=position)
            row += 1
    if image_path is not None:
        img = Image(image_path)
        img.width = _IMAGE_WIDTH_PX
        img.height = _IMAGE_HEIGHT_PX
        img.anchor = _IMAGE_ANCHOR
        ws.add_image(img)
    wb.save(out_path)
    return out_path


def _rows_for(
    roster: SchoolRoster, committee: str, count: int
) -> list[tuple[str, str, str | None]]:
    if _KIND[committee] == "SINGLE":
        return [(roster.name, committee, None)]
    weight = WEIGHT[committee]
    slot_count = count // weight
    countries = roster.countries[:slot_count]
    if weight == 1:
        return [(roster.name, committee, country) for country in countries]
    return [
        (roster.name, committee, f"{country}, {seat}")
        for country in countries
        for seat in (1, 2)
    ]


def _sanitize(name: str) -> str:
    cleaned = "".join("_" if ch in _INVALID_FILENAME_CHARS else ch for ch in name)
    return cleaned.strip()
