"""Load committee capacities from RoomNumbers.xlsx."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from whsmun.committees import XLSX_TO_COMMITTEE


def load_capacities(path: Path) -> dict[str, int]:
    """Return {canonical_committee_name: capacity}.

    Reads the 'Committee Numbers' sheet, columns (dept, committee, count, ...).
    Unknown committee labels raise ValueError listing every offender at once.
    """
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["Committee Numbers"]

    capacities: dict[str, int] = {}
    unknown: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        _dept, committee, count, *_ = row
        if not committee or count is None:
            continue
        canonical = XLSX_TO_COMMITTEE.get(committee)
        if canonical is None:
            unknown.append(committee)
            continue
        capacities[canonical] = int(count)

    if unknown:
        raise ValueError(f"Unknown committee rows in xlsx: {unknown}")
    return capacities
