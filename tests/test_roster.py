"""Roster generator: row layout, CSV round-trip, and end-to-end xlsx output."""
from __future__ import annotations

from importlib.resources import as_file, files
from pathlib import Path

import pytest
from openpyxl import load_workbook

from whsmun.committees import ALL_COMMITTEES
from whsmun.reporting import SchoolRoster, read_assignments_csv, write_rosters
from whsmun.reporting.csv_writer import _HEADER
from whsmun.reporting.roster import _rows_for, _sanitize, _write_one


def _roster(name: str, countries: tuple[str, ...], **placements: int) -> SchoolRoster:
    base = {c: 0 for c in ALL_COMMITTEES}
    base.update(placements)
    return SchoolRoster(
        name=name,
        lottery_country=countries[0] if countries else None,
        countries=countries,
        placements=base,
    )


def test_rows_for_ga_double_del():
    r = _roster("Madison Country Day", ("Brazil", "Turkiye", "Ukraine"))
    rows = _rows_for(r, "SOCHUM", 6)
    assert rows == [
        ("Madison Country Day", "SOCHUM", "Brazil, 1"),
        ("Madison Country Day", "SOCHUM", "Brazil, 2"),
        ("Madison Country Day", "SOCHUM", "Turkiye, 1"),
        ("Madison Country Day", "SOCHUM", "Turkiye, 2"),
        ("Madison Country Day", "SOCHUM", "Ukraine, 1"),
        ("Madison Country Day", "SOCHUM", "Ukraine, 2"),
    ]


def test_rows_for_ga_single_del():
    r = _roster("School", ("Brazil", "Turkiye", "Ukraine"))
    rows = _rows_for(r, "ECOFIN", 3)
    assert rows == [
        ("School", "ECOFIN", "Brazil"),
        ("School", "ECOFIN", "Turkiye"),
        ("School", "ECOFIN", "Ukraine"),
    ]


def test_rows_for_ga_partial_slots_takes_first_n_countries():
    r = _roster("Edgewood", ("Mexico", "Israel", "South Korea"))
    rows = _rows_for(r, "GA Ad-Hoc", 2)
    assert rows == [
        ("Edgewood", "GA Ad-Hoc", "Mexico"),
        ("Edgewood", "GA Ad-Hoc", "Israel"),
    ]


def test_rows_for_double_del_partial_slots():
    r = _roster("Edgewood", ("Mexico", "Israel", "South Korea"))
    rows = _rows_for(r, "SOCHUM", 4)
    assert rows == [
        ("Edgewood", "SOCHUM", "Mexico, 1"),
        ("Edgewood", "SOCHUM", "Mexico, 2"),
        ("Edgewood", "SOCHUM", "Israel, 1"),
        ("Edgewood", "SOCHUM", "Israel, 2"),
    ]


def test_rows_for_single_seat_leaves_position_blank():
    r = _roster("School", ("Brazil",))
    rows = _rows_for(r, "EU", 1)
    assert rows == [("School", "EU", None)]


def test_read_assignments_csv_roundtrip(tmp_path: Path):
    csv_path = tmp_path / "assignments.csv"
    header = _HEADER
    placement_values = {c: 0 for c in ALL_COMMITTEES}
    placement_values["SOCHUM"] = 4
    placement_values["ECOFIN"] = 2
    placement_values["EU"] = 1
    row = [
        "Test School",
        "Brazil",
        "Brazil, Japan",
        2,
        7,
        *(placement_values[c] for c in ALL_COMMITTEES),
        7,
        0,
    ]
    import csv
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerow(row)
    rosters = read_assignments_csv(csv_path)
    assert len(rosters) == 1
    r = rosters[0]
    assert r.name == "Test School"
    assert r.lottery_country == "Brazil"
    assert r.countries == ("Brazil", "Japan")
    assert r.placements["SOCHUM"] == 4
    assert r.placements["ECOFIN"] == 2
    assert r.placements["EU"] == 1
    assert r.placements["AU"] == 0


def test_read_assignments_csv_handles_blank_lottery(tmp_path: Path):
    csv_path = tmp_path / "assignments.csv"
    placement_values = {c: 0 for c in ALL_COMMITTEES}
    row = [
        "No Lottery School",
        "",
        "France, Italy",
        2,
        0,
        *(placement_values[c] for c in ALL_COMMITTEES),
        0,
        0,
    ]
    import csv
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(_HEADER)
        w.writerow(row)
    rosters = read_assignments_csv(csv_path)
    assert rosters[0].lottery_country is None
    assert rosters[0].countries == ("France", "Italy")


def test_filename_sanitization():
    assert _sanitize("Normal School") == "Normal School"
    assert _sanitize("St. Mary's / Holy Cross") == "St. Mary's _ Holy Cross"
    assert _sanitize('Bad:*?"<>|\\Chars') == "Bad________Chars"
    assert _sanitize("  Trim Me  ") == "Trim Me"


def _template_path():
    return as_file(files("whsmun.templates") / "roster_template.xlsx")


def test_write_one_emits_expected_rows(tmp_path: Path):
    r = _roster(
        "Sample",
        ("Brazil", "Turkiye", "Ukraine"),
        SOCHUM=6, DISEC=6, ECOFIN=3, SPECPOL=3, **{"GA Ad-Hoc": 3},
        EU=1, AU=1,
    )
    with _template_path() as template:
        out = _write_one(template, tmp_path, r)
    wb = load_workbook(out)
    ws = wb["ROSTER TEMPLATE"]
    rows = []
    for r_idx in range(7, 100):
        a = ws.cell(row=r_idx, column=1).value
        if a is None:
            break
        rows.append((a, ws.cell(row=r_idx, column=2).value, ws.cell(row=r_idx, column=3).value))
    # 6 SOCHUM + 6 DISEC + 3 ECOFIN + 3 SPECPOL + 3 GA Ad-Hoc + 1 EU + 1 AU = 23
    assert len(rows) == 23
    assert rows[0] == ("Sample", "SOCHUM", "Brazil, 1")
    assert rows[5] == ("Sample", "SOCHUM", "Ukraine, 2")
    assert rows[12] == ("Sample", "ECOFIN", "Brazil")
    assert rows[21] == ("Sample", "EU", None)
    assert rows[22] == ("Sample", "AU", None)


def test_write_rosters_end_to_end(tmp_path: Path):
    rosters = [
        _roster("School A", ("X", "Y"), ECOFIN=2, EU=1),
        _roster("School B", ("Z",), SPECPOL=1, ICJ=1),
    ]
    output_dir = tmp_path / "Rosters"
    with _template_path() as template:
        count = write_rosters(template, output_dir, rosters)
    assert count == 2
    files_out = sorted(p.name for p in output_dir.iterdir())
    assert files_out == [
        "School A WHSMUN Roster.xlsx",
        "School B WHSMUN Roster.xlsx",
    ]
    wb = load_workbook(output_dir / "School A WHSMUN Roster.xlsx")
    ws = wb["ROSTER TEMPLATE"]
    # Header rows untouched.
    assert ws["A1"].value == "WHSMUN LV Committee Assignments"
    assert ws["A6"].value == "School"
    assert ws["B6"].value == "Committee"
    assert ws["C6"].value == "Position"
    # Merged ranges from template preserved (smoke check on formatting).
    assert len(list(ws.merged_cells.ranges)) == 8
    # Data starts at row 7.
    assert ws["A7"].value == "School A"
    assert ws["B7"].value == "ECOFIN"
    assert ws["C7"].value == "X"


def test_write_rosters_wipes_existing_dir(tmp_path: Path):
    output_dir = tmp_path / "Rosters"
    output_dir.mkdir()
    stale = output_dir / "stale.xlsx"
    stale.write_text("stale")
    with _template_path() as template:
        write_rosters(template, output_dir, [_roster("School", ("X",), EU=1)])
    assert not stale.exists()
    assert (output_dir / "School WHSMUN Roster.xlsx").exists()
