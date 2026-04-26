"""Load registration CSV, room capacity xlsx, and lottery JSON."""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from committees import CSV_LABEL_TO_COMMITTEE, XLSX_TO_COMMITTEE
from models import School


def _normalize_school(name: str) -> str:
    """Lowercase + strip non-alphanumerics so 'Madison Country Day' and
    'madison country day' compare equal."""
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def _first_int(text: str) -> int:
    m = re.search(r"-?\d+", text or "")
    if not m:
        raise ValueError(f"no integer in {text!r}")
    return int(m.group())


def load_lottery(path: Path) -> dict[str, str]:
    """Returns {normalized_school_name: country}."""
    raw = json.loads(path.read_text())
    return {_normalize_school(school): country for school, country in raw.items()}


def load_capacities(path: Path) -> dict[str, int]:
    """Returns {canonical_committee_name: capacity}."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Committee Numbers"]
    caps: dict[str, int] = {}
    unknown: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        _dept, committee, count, *_ = row
        if not committee or count is None:
            continue
        canonical = XLSX_TO_COMMITTEE.get(committee)
        if canonical is None:
            unknown.append(committee)
            continue
        caps[canonical] = int(count)
    if unknown:
        raise ValueError(f"Unknown committee rows in xlsx: {unknown}")
    return caps


def _build_committee_column_map(headers: list[str]) -> dict[str, str]:
    """Map each CSV committee-question column header to a canonical committee
    name, using the trailing identifier in CSV_LABEL_TO_COMMITTEE."""
    mapping: dict[str, str] = {}
    for header in headers:
        if "send a delegate to" not in header.lower():
            continue
        # Try labels longest-first so "JCC Side A" wins over a hypothetical "JCC".
        for label in sorted(CSV_LABEL_TO_COMMITTEE, key=len, reverse=True):
            if label.lower() in header.lower():
                mapping[header] = CSV_LABEL_TO_COMMITTEE[label]
                break
        else:
            raise ValueError(f"CSV committee column not recognized: {header!r}")
    return mapping


def load_schools(csv_path: Path, lottery: dict[str, str]) -> list[School]:
    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        return []

    headers = list(rows[0].keys())
    committee_columns = _build_committee_column_map(headers)
    size_key = next(k for k in headers if k.startswith("How large is your Delegation"))
    country_key = next(k for k in headers if k.startswith("Besides the country you drew"))

    schools: list[School] = []
    for idx, row in enumerate(rows):
        name = (row.get("Name of School") or "").strip()
        if not name:
            raise ValueError(f"row {idx}: missing 'Name of School'")
        total = _first_int(row[size_key])
        extras = _first_int(row[country_key]) if (row[country_key] or "").strip() else 0
        lottery_country = lottery.get(_normalize_school(name))
        yes_committees = tuple(
            committee
            for header, committee in committee_columns.items()
            if (row.get(header) or "").strip().lower() == "yes"
        )
        schools.append(
            School(
                row_index=idx,
                name=name,
                total_delegates=total,
                extra_countries=extras,
                lottery_country=lottery_country,
                yes_committees=yes_committees,
            )
        )
    return schools